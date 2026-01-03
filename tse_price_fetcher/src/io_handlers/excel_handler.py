"""
Excel ファイルの読み書きモジュール
"""
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler:
    """Excel ファイルの入出力を行うクラス"""

    def __init__(self, file_path: str):
        """
        Args:
            file_path: Excel ファイルのパス
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None
        self.sheet: Optional[Worksheet] = None

    def read(self, sheet_name: str = None) -> List[Dict]:
        """
        Excel ファイルを読み込み、辞書のリストとして返す

        Args:
            sheet_name: シート名（Noneの場合は最初のシート）

        Returns:
            各行を辞書として格納したリスト
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"ファイルが見つかりません: {self.file_path}")

        self.workbook = load_workbook(self.file_path)

        # シート選択
        if sheet_name:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.active

        # ヘッダー行を取得
        rows = list(self.sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = rows[0]
        data = []

        # データ行を辞書に変換
        for row in rows[1:]:
            row_dict = {}
            for idx, header in enumerate(headers):
                value = row[idx] if idx < len(row) else None
                row_dict[header] = value
            data.append(row_dict)

        return data

    def write(self, data: List[Dict], sheet_name: str = None):
        """
        データを Excel ファイルに書き込む

        Args:
            data: 書き込むデータ（辞書のリスト）
            sheet_name: シート名（Noneの場合は最初のシート）
        """
        if not data:
            return

        # 既存ファイルの読み込み、または新規作成
        if self.file_path.exists():
            self.workbook = load_workbook(self.file_path)
            if sheet_name:
                if sheet_name in self.workbook.sheetnames:
                    self.sheet = self.workbook[sheet_name]
                else:
                    self.sheet = self.workbook.create_sheet(sheet_name)
            else:
                self.sheet = self.workbook.active
        else:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            if sheet_name:
                self.sheet.title = sheet_name

        # シートをクリア
        self.sheet.delete_rows(1, self.sheet.max_row)

        # ヘッダーを決定的な順序で書き込む
        # すべての行から全キーを収集し、ソート
        all_keys = set()
        for row in data:
            all_keys.update(row.keys())

        # 決定的な順序: tyo.code, base_date を最初に、その後アルファベット順
        priority_keys = ['tyo.code', 'base_date', 'adjusted_base_date', 'fetch_status']
        headers = []

        for key in priority_keys:
            if key in all_keys:
                headers.append(key)
                all_keys.remove(key)

        # 残りのキーをソート
        remaining_keys = sorted(all_keys)
        headers.extend(remaining_keys)

        # ヘッダー行を書き込み
        self.sheet.append(headers)

        # データ行を書き込み
        for row in data:
            row_values = [row.get(header) for header in headers]
            self.sheet.append(row_values)

        # ファイルを保存
        self.workbook.save(self.file_path)

    def close(self):
        """ワークブックを閉じる"""
        if self.workbook:
            self.workbook.close()
