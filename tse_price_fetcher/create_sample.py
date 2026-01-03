#!/usr/bin/env python3
"""
サンプル Excel ファイルを作成するスクリプト
"""
try:
    from openpyxl import Workbook
    from datetime import datetime

    # ワークブックを作成
    wb = Workbook()
    ws = wb.active
    ws.title = "株価データ"

    # ヘッダー
    ws['A1'] = 'tyo.code'
    ws['B1'] = 'base_date'

    # サンプルデータ
    sample_data = [
        ('7203', '2024-01-15'),  # トヨタ自動車
        ('9984', '2024-01-15'),  # ソフトバンクグループ
        ('6758', '2024-01-20'),  # ソニーグループ
        ('4063', '2024-01-25'),  # 信越化学工業
        ('8306', '2024-02-01'),  # 三菱UFJフィナンシャル・グループ
    ]

    # データを書き込み
    for i, (code, date) in enumerate(sample_data, start=2):
        ws[f'A{i}'] = code
        ws[f'B{i}'] = date

    # 列幅を調整
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15

    # 保存
    wb.save('sample_input.xlsx')
    print("✓ sample_input.xlsx を作成しました")

except ImportError:
    print("openpyxl がインストールされていません。")
    print("先に setup.sh を実行してください。")
except Exception as e:
    print(f"エラー: {e}")
